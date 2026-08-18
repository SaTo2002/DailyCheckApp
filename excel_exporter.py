import os
import re

import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image as PILImage
from PIL import ImageOps

# Exact template paths for each area/branch
EXCEL_TEMPLATES = {
    "park": os.path.join(os.path.dirname(__file__), "Exsl", "DailyCheck_Park.xlsx"),
    "kids": os.path.join(os.path.dirname(__file__), "Exsl", "DailyCheck_Kids.xlsx"),
    "kickstrez": os.path.join(
        os.path.dirname(__file__), "Exsl", "DailyCheck_Kickerz.xlsx"
    ),
    "bowling": os.path.join(
        os.path.dirname(__file__), "Exsl", "DailyCheck_Bowling.xlsx"
    ),
}

# Hardcoded dictionaries removed. Using dynamic tags.


def _cell_area_px(ws, fc, fr, tc, tr):
    """
    Calculate the pixel dimensions of a merged cell area (fc:fr → tc:tr, 1-indexed).
    Column width (chars) → px  : (width * 7 + 5)
    Row height   (pts)   → px  : height * 96 / 72
    """
    w = sum(
        int((ws.column_dimensions[get_column_letter(c)].width or 8) * 7 + 5)
        for c in range(fc, tc + 1)
    )
    h = sum(
        int((ws.row_dimensions[r].height or 15) * 96 / 72) for r in range(fr, tr + 1)
    )
    return max(w, 1), max(h, 1)


def export_report_to_excel(
    report_session_id,
    monitor_name,
    area_name,
    checks_dict,
    game_notes_dict,
    game_maps_dict,
    date_str,
    output_xlsx_path,
    orientation="portrait",
):
    """
    Fills local Excel template file directly using openpyxl, populates values and issue map images,
    and saves to output_xlsx_path. Completely replaces Google Sheets dependency.
    """
    # Select template based on area_name
    area_clean = area_name.lower().strip()
    template_path = EXCEL_TEMPLATES.get("park")
    for key, path in EXCEL_TEMPLATES.items():
        if key in area_clean:
            template_path = path
            break

    if not os.path.exists(template_path):
        # Fallback to Park template
        template_path = EXCEL_TEMPLATES["park"]

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Header (Name and Date) will be filled dynamically during the tag scan

    # 2. Map questions rows per Section in Column B
    # Each section (e.g. F.O, Lounge, Airtrack...) maps its own question text -> row number
    section_map = {}
    current_sec = "General"
    section_map[current_sec] = {}
    section_title_rows = set()

    for r in range(1, ws.max_row + 1):
        col_g = str(ws.cell(row=r, column=7).value or "").strip()
        col_h = str(ws.cell(row=r, column=8).value or "").strip()
        cell_val = ws.cell(row=r, column=2).value

        if col_g == "OK" and col_h == "NOK":
            section_title_rows.add(r)
            current_sec = " ".join(str(cell_val).strip().split()).lower()
            section_map[current_sec] = {}
        elif cell_val:
            clean_txt = " ".join(str(cell_val).strip().split()).lower()
            section_map[current_sec][clean_txt] = r

    # 2.5 Scan sheet for Tags ([MAP:*] and [NOTE:*])
    dynamic_map_cells = {}
    dynamic_note_cells = {}

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val.startswith("[MAP:") and val.endswith("]"):
                    game_tag = val[5:-1].strip().lower()

                    merged_range = None
                    for merged in ws.merged_cells.ranges:
                        if cell.coordinate in merged:
                            merged_range = merged
                            break

                    if merged_range:
                        fc = get_column_letter(merged_range.min_col)
                        fr = merged_range.min_row
                        tc = get_column_letter(merged_range.max_col)
                        tr = merged_range.max_row
                        dynamic_map_cells[game_tag] = (f"{fc}{fr}", f"{tc}{tr}")
                    else:
                        dynamic_map_cells[game_tag] = (cell.coordinate, cell.coordinate)

                    cell.value = ""  # Clear the tag

                elif val.startswith("[NOTE:") and val.endswith("]"):
                    game_tag = val[6:-1].strip().lower()
                    dynamic_note_cells[game_tag] = cell.coordinate
                    cell.value = ""  # Clear the tag

                elif "[name]" in val.lower() or "[date]" in val.lower():
                    new_val = re.sub(
                        r"\[name\]", monitor_name, val, flags=re.IGNORECASE
                    )
                    new_val = re.sub(
                        r"\[date\]", date_str, new_val, flags=re.IGNORECASE
                    )
                    cell.value = new_val

    # 3. Fill Checkmarks (Col G: OK, Col H: NOK) per exact section
    for game_name, game_checks in checks_dict.items():
        gn_clean = " ".join(game_name.strip().split()).lower()

        # Match game to Excel section
        matched_sec_dict = None
        for sec_name, q_row_dict in section_map.items():
            if sec_name in gn_clean or gn_clean in sec_name:
                matched_sec_dict = q_row_dict
                break

        # If no section match, fallback to global lookup
        if not matched_sec_dict:
            global_qs = {}
            for qd in section_map.values():
                global_qs.update(qd)
            matched_sec_dict = global_qs

        for q_name, status in game_checks.items():
            clean_q = " ".join(q_name.strip().split()).lower()
            if clean_q in matched_sec_dict:
                r_idx = matched_sec_dict[clean_q]
                if r_idx not in section_title_rows:
                    if status == "OK":
                        ws.cell(row=r_idx, column=7, value="✔")
                    elif status == "NOK":
                        ws.cell(row=r_idx, column=8, value="✔")

    # 4. Fill Comments
    cell_notes_map = {}
    for game_name, note_text in game_notes_dict.items():
        if note_text and note_text.strip():
            target_cell = None
            gn_clean = " ".join(game_name.strip().split()).lower()
            if gn_clean in dynamic_note_cells:
                target_cell = dynamic_note_cells[gn_clean]
            else:
                for tag_game, cell_addr in dynamic_note_cells.items():
                    if tag_game in gn_clean or gn_clean in tag_game:
                        target_cell = cell_addr
                        break
            if target_cell:
                if target_cell not in cell_notes_map:
                    cell_notes_map[target_cell] = []
                cell_notes_map[target_cell].append(note_text.strip())

    from openpyxl.styles import Alignment, Font

    protected_rows = set()
    for from_cell, to_cell in dynamic_map_cells.values():
        fr = int("".join(filter(str.isdigit, from_cell)))
        tr = int("".join(filter(str.isdigit, to_cell)))
        for r in range(fr, tr + 1):
            protected_rows.add(r)

    for cell_addr in set(dynamic_note_cells.values()):
        notes_list = cell_notes_map.get(cell_addr, [])
        cell = ws[cell_addr]

        # Check if we should hide this note section
        is_na_or_empty = True
        for n in notes_list:
            if n and n.strip().upper() != "N/A":
                is_na_or_empty = False
                break

        if is_na_or_empty:
            merged_range = None
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    merged_range = merged
                    break
            
            is_visible = False
            # Hide the row(s) only if they don't intersect with map placeholders
            if merged_range:
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    if r not in protected_rows:
                        ws.row_dimensions[r].hidden = True
                    else:
                        is_visible = True
            else:
                if cell.row not in protected_rows:
                    ws.row_dimensions[cell.row].hidden = True
                else:
                    is_visible = True
            
            if is_visible:
                cell.value = "N/A"
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                cell.font = Font(name="Arial", size=11, bold=True, color="888888")
            else:
                cell.value = ""

            continue

        # If there are valid notes, write them normally
        all_paragraphs = []
        for n in notes_list:
            if not n or n.strip().upper() == "N/A":
                continue
            lines = [line.strip() for line in n.splitlines() if line.strip()]
            if lines:
                # Add a bullet point to clearly separate each distinct item/issue
                paragraph = " • " + "   • ".join(lines)
                all_paragraphs.append(paragraph)

        cell.value = "   ".join(all_paragraphs)

        # Word-like Rich Text Alignment & Styling (RTL Arabic, Wrap Text, Top Alignment, Clean Font)
        cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="right", readingOrder=2
        )
        cell.font = Font(name="Arial", size=10, bold=False)

    created_temp_files = []
    try:
        # 5. Insert Map Images stretched to fill merged cells exactly (TwoCellAnchor)
        if game_maps_dict:
            for game_name, map_info in game_maps_dict.items():
                drawing_path = (
                    map_info.get("drawing", "")
                    if isinstance(map_info, dict)
                    else map_info
                )
                base_path = (
                    map_info.get("base", "") if isinstance(map_info, dict) else ""
                )

                if drawing_path or base_path:
                    d_path = drawing_path.lstrip("/") if drawing_path else ""
                    b_path = base_path.lstrip("/") if base_path else ""

                    if d_path and not os.path.isabs(d_path):
                        d_path = os.path.abspath(d_path)
                    if b_path and not os.path.isabs(b_path):
                        b_path = os.path.abspath(b_path)

                    img_to_use = (
                        d_path if (d_path and os.path.exists(d_path)) else b_path
                    )

                    if img_to_use and os.path.exists(img_to_use):
                        cell_range = None
                        gn_clean = (
                            game_name.strip().lower().replace("(map)", "").strip()
                        )

                        if gn_clean in dynamic_map_cells:
                            cell_range = dynamic_map_cells[gn_clean]
                        else:
                            for tag_game, cell_tuple in dynamic_map_cells.items():
                                if tag_game in gn_clean or gn_clean in tag_game:
                                    cell_range = cell_tuple
                                    break

                        if cell_range:
                            try:
                                from_cell, to_cell = cell_range

                                # 1. مسح أي قيمة نصية قديمة في الخلية الأولى
                                ws[from_cell].value = None

                                # 2. إزالة الصور القديمة المحطوطة فوق نفس المنطقة في القالب
                                fc = column_index_from_string(
                                    "".join(filter(str.isalpha, from_cell))
                                )
                                fr = int("".join(filter(str.isdigit, from_cell)))
                                tc = column_index_from_string(
                                    "".join(filter(str.isalpha, to_cell))
                                )
                                tr = int("".join(filter(str.isdigit, to_cell)))

                                imgs_to_remove = [
                                    im
                                    for im in ws._images
                                    if hasattr(im, "anchor")
                                    and hasattr(im.anchor, "_from")
                                    and fr - 1 <= im.anchor._from.row <= tr
                                    and fc - 1 <= im.anchor._from.col <= tc
                                ]
                                for old_im in imgs_to_remove:
                                    if old_im in ws._images:
                                        ws._images.remove(old_im)

                                # 3. دمج الرسمة مع الخريطة الأصلية لو الاثنين موجودين
                                if (
                                    d_path
                                    and os.path.exists(d_path)
                                    and b_path
                                    and os.path.exists(b_path)
                                ):
                                    base_img = PILImage.open(b_path).convert("RGBA")
                                    draw_img = PILImage.open(d_path).convert("RGBA")
                                    draw_rszd = draw_img.resize(
                                        base_img.size, PILImage.Resampling.LANCZOS
                                    )
                                    final_img = PILImage.alpha_composite(
                                        base_img, draw_rszd
                                    )
                                    temp_dir = os.path.join(
                                        "static", "uploads", "temp_composites"
                                    )
                                    os.makedirs(temp_dir, exist_ok=True)
                                    img_to_insert = os.path.abspath(
                                        os.path.join(
                                            temp_dir,
                                            f"excel_{os.path.basename(img_to_use)}",
                                        )
                                    )
                                    final_img.save(img_to_insert, "PNG")
                                    created_temp_files.append(img_to_insert)
                                else:
                                    img_to_insert = img_to_use

                                # 4. Contain-fit بدون أي تعديلات على الاتجاه
                                cell_w, cell_h = _cell_area_px(ws, fc, fr, tc, tr)

                                src = PILImage.open(img_to_insert)
                                src_w, src_h = src.size

                                # حساب نسبة التصغير للحفاظ على الـ Aspect Ratio بدون Stretch
                                scale = min(cell_w / src_w, cell_h / src_h)
                                fit_w = int(src_w * scale)
                                fit_h = int(src_h * scale)

                                img_obj = OpenPyxlImage(img_to_insert)
                                img_obj.width = fit_w
                                img_obj.height = fit_h
                                ws.add_image(img_obj, from_cell)

                            except Exception as img_err:
                                print(
                                    f"Error adding image to Excel ({game_name}): {img_err}"
                                )

        # Set Page Setup Properties: Dynamic Orientation & Fit All Columns on One Page
        if str(orientation).lower() == "landscape":
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        else:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Ensure target output directory exists
        os.makedirs(os.path.dirname(output_xlsx_path), exist_ok=True)
        wb.save(output_xlsx_path)
        return output_xlsx_path
    finally:
        # Clean up temporary composite image files after saving Excel report
        for temp_file in created_temp_files:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception as clean_err:
                    print(
                        f"Warning: Could not remove temporary file {temp_file}: {clean_err}"
                    )
